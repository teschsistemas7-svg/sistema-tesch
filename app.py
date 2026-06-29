import os
import psycopg2
import pandas as pd
from flask import Flask, render_template, request, redirect, session
from datetime import datetime
from io import BytesIO
from werkzeug.security import generate_password_hash, check_password_hash





app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev_key")

# ======================
# BASE DE DATOS
# ======================




def get_db():

    DATABASE_URL = os.environ.get("DATABASE_URL")

    # ✅ Si estás en Render usa esa
    if DATABASE_URL:
        return psycopg2.connect(DATABASE_URL, sslmode='require')

    # ✅ Si estás en local usa la externa
    return psycopg2.connect(
        "postgresql://tesch:LdYWnIgIMZ5E5ngP1HASUC4LzhyYcqut@dpg-d90nhr0js32c73dcc72g-a.ohio-postgres.render.com/tesch",
        sslmode='require'
    )

def init_db():
    conn = get_db()
    cursor = conn.cursor()

    
    


    cursor.execute('''CREATE TABLE IF NOT EXISTS usuarios (
        usuario TEXT PRIMARY KEY,
        password TEXT,
        rol TEXT,
        cambio_password INTEGER DEFAULT 0)''')

    cursor.execute('''CREATE TABLE IF NOT EXISTS alumnos (
        matricula TEXT PRIMARY KEY,
        nombre TEXT,
        carrera TEXT)''')

    
    cursor.execute('''CREATE TABLE IF NOT EXISTS actividades (
    id SERIAL PRIMARY KEY,
    nombre TEXT,
    docente TEXT
    )''')


    cursor.execute('''CREATE TABLE IF NOT EXISTS periodos (
        id SERIAL PRIMARY KEY,
        periodo TEXT,
        activo INTEGER)''')

    cursor.execute('''CREATE TABLE IF NOT EXISTS inscripciones (
        id SERIAL PRIMARY KEY,
        matricula TEXT,
        semestre TEXT,
        genero TEXT,
        telefono TEXT,
        actividad TEXT,
        periodo TEXT)''')
    cursor.execute("""
    CREATE UNIQUE INDEX IF NOT EXISTS idx_inscripcion
    ON inscripciones(matricula, periodo)
    """)
    
    
    cursor.execute('''CREATE TABLE IF NOT EXISTS control_fechas (
        id SERIAL PRIMARY KEY,
        fecha_inicio TEXT,
        fecha_fin TEXT
    )''')

    cursor.execute('''CREATE TABLE IF NOT EXISTS eventos (
    id SERIAL PRIMARY KEY,
    docente TEXT,
    actividad TEXT,
    nombre_evento TEXT,
    institucion TEXT,
    fecha TEXT,
    participantes INTEGER,
    mujeres INTEGER,
    hombres INTEGER,
    resultados TEXT,
    periodo TEXT
    )''')

    
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS resultados (
    id SERIAL PRIMARY KEY,
    matricula TEXT,
    actividad TEXT,
    periodo TEXT,
    resultado TEXT
    )
    ''')

    
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS evaluaciones (
    id SERIAL PRIMARY KEY,
    matricula TEXT,
    actividad TEXT,
    periodo TEXT,
    c1 INTEGER,
    c2 INTEGER,
    c3 INTEGER,
    c4 INTEGER,
    c5 INTEGER,
    c6 INTEGER,
    c7 INTEGER,
    promedio REAL,
    nivel TEXT,
    observaciones TEXT
    )
    ''')




    conn.commit()
    conn.close()



from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from io import BytesIO




# ======================
# Registrar estudiante
# ======================


@app.route('/crear_alumno', methods=['GET','POST'])
def crear_alumno():

    if 'usuario' not in session:
        return redirect('/')

    if session.get('rol') != 'admin':
        return "Acceso denegado"

    mensaje = ""

    if request.method == 'POST':

        # ✅ limpiar datos
        matricula = request.form.get('matricula', '').strip()
        nombre = request.form.get('nombre', '').strip()
        carrera = request.form.get('carrera', '').strip()

        # ✅ validar campos
        if not matricula or not nombre:
            mensaje = "❌ Debes llenar matrícula y nombre"
            return render_template("crear_alumno.html", mensaje=mensaje)

        conn = get_db()
        cursor = conn.cursor()

        try:
            # ✅ validar si ya existe en alumnos
            cursor.execute("SELECT 1 FROM alumnos WHERE matricula=%s", (matricula,))
            if cursor.fetchone():
                conn.close()
                mensaje = "❌ La matrícula ya existe"
                return render_template("crear_alumno.html", mensaje=mensaje)

            # ✅ validar si ya existe como usuario
            cursor.execute("SELECT 1 FROM usuarios WHERE usuario=%s", (matricula,))
            if cursor.fetchone():
                conn.close()
                mensaje = "❌ Ya existe un usuario con esa matrícula"
                return render_template("crear_alumno.html", mensaje=mensaje)

            # ✅ insertar alumno
            cursor.execute("""
            INSERT INTO alumnos (matricula, nombre, carrera)
            VALUES (%s, %s, %s)
            """, (matricula, nombre, carrera))

            # ✅ crear usuario automático
            password_hash = generate_password_hash(matricula)

            cursor.execute("""
            INSERT INTO usuarios (usuario, password, rol, cambio_password)
            VALUES (%s, %s, 'alumno', 0)
            """, (matricula, password_hash))

            conn.commit()
            conn.close()

            return redirect('/admin_alumnos')

        except Exception as e:
            conn.close()
            return f"❌ Error: {str(e)}"

    return render_template("crear_alumno.html", mensaje=mensaje)



# ======================
# Exportar asistencia excel
# ======================

@app.route('/exportar_asistencia_excel')
def exportar_asistencia_excel():

    if session.get('rol') != 'docente':
        return "Acceso denegado"

    conn = get_db()
    cursor = conn.cursor()

    docente = session['usuario']
    periodo_activo = obtener_periodo_activo()

    # ✅ obtener taller desde la URL
    taller = request.args.get('taller')

    if not taller:
        conn.close()
        return "❌ Debes seleccionar un taller"

    # ✅ NUEVO 🔥 OBTENER DÍAS SELECCIONADOS
    dias = request.args.getlist('dias')

    if not dias:
        conn.close()
        return "❌ Selecciona al menos un día"

    cursor.execute("""
    SELECT a.matricula, a.nombre, a.carrera,
           i.semestre, i.actividad, i.periodo
    FROM alumnos a
    JOIN inscripciones i ON a.matricula=i.matricula
    JOIN actividades act ON act.nombre=i.actividad
    WHERE act.docente=%s 
      AND i.periodo=%s 
      AND i.actividad=%s
    """, (docente, periodo_activo, taller))

    alumnos = cursor.fetchall()
    conn.close()

    if not alumnos:
        return "❌ No hay alumnos en este taller"

    actividad = alumnos[0][4]
    periodo = alumnos[0][5]

    def abreviar(carrera):
        mapa = {
            "Ingeniería en Sistemas Computacionales": "ISC",
            "Ingeniería en Energías Renovables": "IER",
            "Ingeniería en Desarrollo Comunitario": "IDC",
            "Ingeniería en Gestión Empresarial": "IGE",
            "Ingeniería en Semiconductores": "IS"
        }
        return mapa.get(carrera, carrera)

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    center = Alignment(horizontal="center", vertical="center")
    bold = Font(bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # ✅ ENCABEZADO
    ws.merge_cells("A1:Z1")
    ws["A1"] = "TECNOLÓGICO DE ESTUDIOS SUPERIORES DE CHICOLOAPAN"
    ws["A1"].alignment = center
    ws["A1"].font = bold

    ws["B3"] = "Taller de:"
    ws["C3"] = actividad

    ws["B4"] = "Docente:"
    ws["C4"] = docente

    ws["B5"] = "Parcial:"
    ws["C5"] = "_____   (1°, 2° o 3°)"

    ws["B6"] = "Periodo:"
    ws["C6"] = periodo

    # ✅ TABLA
    fila_inicio = 7

    headers = ["No", "Matrícula", "Nombre del estudiante", "Carrera", "Semestre"]

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=fila_inicio, column=col, value=h)
        cell.font = bold
        cell.alignment = center
        cell.border = border

    # ✅ COLUMNAS DINÁMICAS 🔥
    start_col = 6
    col = start_col

    total_semanas = 16

    for semana in range(total_semanas):
        for dia in dias:

            cell_fecha = ws.cell(row=fila_inicio, column=col, value=dia)
            cell_fecha.alignment = Alignment(
                horizontal="center",
                vertical="center",
                textRotation=90
            )
            cell_fecha.border = border

            cell_dia = ws.cell(row=fila_inicio + 1, column=col, value=f"S{semana+1}")
            cell_dia.alignment = center
            cell_dia.border = border

            col += 1

    # ✅ ALUMNOS
    for i, a in enumerate(alumnos, start=1):

        fila = fila_inicio + 2 + i

        ws.cell(row=fila, column=1, value=i)
        ws.cell(row=fila, column=2, value=a[0])
        ws.cell(row=fila, column=3, value=a[1])
        ws.cell(row=fila, column=4, value=abreviar(a[2]))
        ws.cell(row=fila, column=5, value=a[3])

        for j in range(col - start_col):
            cell = ws.cell(row=fila, column=start_col + j)
            cell.border = border

    # ✅ AJUSTES
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 6

    # ✅ AJUSTE DE IMPRESIÓN 🔥
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name=f"lista_asistencia_{actividad}.xlsx",
        as_attachment=True
    )





@app.route('/admin_eliminar_inscripcion/<matricula>')
def admin_eliminar_inscripcion(matricula):

    if 'usuario' not in session:
        return redirect('/')

    if session.get('rol') != 'admin':
        return "Acceso denegado"

    conn = get_db()
    cursor = conn.cursor()

    # ✅ periodo activo
    cursor.execute("SELECT periodo FROM periodos WHERE activo=1")
    dato = cursor.fetchone()
    periodo = dato[0] if dato else ""

    # ✅ eliminar inscripción
    cursor.execute("""
    DELETE FROM inscripciones
    WHERE matricula=%s AND periodo=%s
    """, (matricula, periodo))

    # ✅ eliminar evaluación y resultado
    cursor.execute("""
    DELETE FROM resultados
    WHERE matricula=%s AND periodo=%s
    """, (matricula, periodo))

    cursor.execute("""
    DELETE FROM evaluaciones
    WHERE matricula=%s AND periodo=%s
    """, (matricula, periodo))

    conn.commit()
    conn.close()

    return redirect('/admin_alumnos')





@app.route('/limpiar_duplicados')
def limpiar_duplicados():

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    DELETE FROM inscripciones
    WHERE id NOT IN (
        SELECT MIN(id)
        FROM inscripciones
        GROUP BY matricula, periodo
    );
    """)

    conn.commit()
    conn.close()

    return ""


# ======================
# obtener periodo activo
# ======================

def obtener_periodo_activo():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT periodo FROM periodos WHERE activo=1")
    dato = cursor.fetchone()
    periodo = dato[0] if dato else ""
    conn.close()
    return periodo



# ======================
# Lista asistencia
# ======================

@app.route('/lista_asistencia')
def lista_asistencia():

    if session.get('rol') != 'docente':
        return "Acceso denegado"

    conn = get_db()
    cursor = conn.cursor()

    docente = session['usuario']
    periodo_activo = obtener_periodo_activo()

    # ✅ obtener parámetros
    taller = request.args.get('taller')
    dias = request.args.getlist('dias')  # 🔥 AQUÍ VA LO QUE FALTABA

    if not taller:
        conn.close()
        return "❌ Debes seleccionar un taller"

    if not dias:
        conn.close()
        return "❌ Selecciona al menos un día"

    # ✅ VALIDAR QUE EL TALLER ES DEL DOCENTE
    cursor.execute("""
    SELECT 1 FROM actividades
    WHERE nombre=%s AND docente=%s
    """, (taller, docente))

    if not cursor.fetchone():
        conn.close()
        return "❌ No tienes acceso a este taller"

    # ✅ CONSULTA
    cursor.execute("""
    SELECT a.matricula, a.nombre, a.carrera,
           i.semestre, i.telefono, i.actividad, i.periodo
    FROM alumnos a
    JOIN inscripciones i ON a.matricula = i.matricula
    JOIN actividades act ON act.nombre = i.actividad
    WHERE act.docente = %s 
    AND i.periodo = %s
    AND i.actividad = %s
    ORDER BY a.nombre ASC
    """, (docente, periodo_activo, taller))

    alumnos = cursor.fetchall()
    conn.close()

    if not alumnos:
        return "❌ No hay alumnos en este taller"

    actividad = alumnos[0][5]
    periodo = alumnos[0][6]

    return render_template(
        'lista_asistencia.html',
        alumnos=alumnos,
        docente=docente,
        actividad=actividad,
        periodo=periodo,
        taller=taller,
        dias=dias   # 🔥 IMPORTANTE
    )
# ======================
# Editar alumnos incritos por dicente
# ======================



@app.route('/editar_alumno_docente/<matricula>', methods=['GET','POST'])
def editar_alumno_docente(matricula):

    if session.get('rol') != 'docente':
        return "Acceso denegado"

    conn = get_db()
    cursor = conn.cursor()

    docente = session['usuario']

    
    periodo_activo = obtener_periodo_activo()

    cursor.execute("""
    SELECT i.actividad
    FROM inscripciones i
    JOIN actividades a ON a.nombre = i.actividad
    WHERE i.matricula=%s AND a.docente=%s AND i.periodo=%s
    """, (matricula, docente, periodo_activo))


    if not cursor.fetchone():
        conn.close()
        return "❌ Este alumno no pertenece a tu taller"

    if request.method == 'POST':

        try:
            nombre = request.form['nombre']
            carrera = request.form['carrera']
            semestre = request.form['semestre']
            telefono = request.form['telefono']

            # ✅ actualizar alumno
            cursor.execute("""
            UPDATE alumnos
            SET nombre=%s, carrera=%s
            WHERE matricula=%s
            """, (nombre, carrera, matricula))

            # ✅ actualizar inscripción
            
            cursor.execute("""
            UPDATE inscripciones
            SET semestre=%s, telefono=%s
            WHERE matricula=%s AND periodo=%s
            """, (semestre, telefono, matricula, periodo_activo))


            conn.commit()
            conn.close()

            return redirect('/docente')

        except Exception as e:
            conn.close()
            return f"❌ Error: {str(e)}"

    # ✅ obtener datos
    
    cursor.execute("""
    SELECT a.matricula, a.nombre, a.carrera,
       i.semestre, i.telefono, i.actividad
    FROM alumnos a
    JOIN inscripciones i ON a.matricula=i.matricula
    WHERE a.matricula=%s AND i.periodo=%s
    """, (matricula, periodo_activo))


    alumno = cursor.fetchone()

    conn.close()

    return render_template("editar_alumno_docente.html", alumno=alumno)



# ======================
# Editar alumnos incritos-------------------------------------
# ======================

@app.route('/editar_inscripcion_admin/<matricula>', methods=['GET','POST'])
def editar_inscripcion_admin(matricula):

    if session.get('rol') != 'admin':
        return "Acceso denegado"

    conn = get_db()
    cursor = conn.cursor()

    mensaje = ""

    if request.method == 'POST':

        nueva_matricula = request.form['matricula'].strip()
        nombre = request.form['nombre']
        carrera = request.form['carrera']
        semestre = request.form['semestre']
        genero = request.form['genero']
        telefono = request.form['telefono']
        actividad = request.form['actividad']

        try:
            # ✅ VALIDAR SI CAMBIÓ MATRÍCULA
            if nueva_matricula != matricula:

                cursor.execute("SELECT * FROM alumnos WHERE matricula=%s", (nueva_matricula,))
                if cursor.fetchone():
                    mensaje = "❌ La matrícula ya existe. No se puede duplicar."
                    
                    # recargar datos
                    
                    periodo_activo = obtener_periodo_activo()

                    cursor.execute("""
                    SELECT a.matricula, a.nombre, a.carrera,
                        i.semestre, i.genero, i.telefono, i.actividad
                    FROM alumnos a
                    LEFT JOIN inscripciones i ON a.matricula=i.matricula
                    WHERE a.matricula=%s AND i.periodo=%s
                    """, (matricula, periodo_activo))

                    
                    alumno = cursor.fetchone()

                    cursor.execute("SELECT nombre FROM actividades")
                    talleres = cursor.fetchall()

                    conn.close()

                    return render_template(
                        "editar_inscripcion_admin.html",
                        alumno=alumno,
                        talleres=talleres,
                        mensaje=mensaje
                    )

                # ✅ ACTUALIZAR EN TODAS LAS TABLAS 🔥
                cursor.execute("UPDATE alumnos SET matricula=%s WHERE matricula=%s", (nueva_matricula, matricula))
                cursor.execute("UPDATE usuarios SET usuario=%s WHERE usuario=%s", (nueva_matricula, matricula))
                cursor.execute("UPDATE inscripciones SET matricula=%s WHERE matricula=%s", (nueva_matricula, matricula))
                cursor.execute("UPDATE resultados SET matricula=%s WHERE matricula=%s", (nueva_matricula, matricula))
                cursor.execute("UPDATE evaluaciones SET matricula=%s WHERE matricula=%s", (nueva_matricula, matricula))

                matricula = nueva_matricula  # actualizar referencia

            # ✅ ACTUALIZAR DATOS
            cursor.execute("""
            UPDATE alumnos
            SET nombre=%s, carrera=%s
            WHERE matricula=%s
            """, (nombre, carrera, matricula))

            periodo_activo = obtener_periodo_activo()
            cursor.execute("""
            UPDATE inscripciones
            SET semestre=%s, genero=%s, telefono=%s, actividad=%s
            WHERE matricula=%s AND periodo=%s
            """, (semestre, genero, telefono, actividad, matricula, periodo_activo))


            conn.commit()
            conn.close()

            return redirect('/admin_alumnos')

        except Exception as e:
            conn.close()
            return f"❌ Error: {str(e)}"

    # ✅ GET
    
    periodo_activo = obtener_periodo_activo()

    cursor.execute("""
    SELECT a.matricula, a.nombre, a.carrera,
       i.semestre, i.genero, i.telefono, i.actividad
    FROM alumnos a
    LEFT JOIN inscripciones i ON a.matricula=i.matricula
    WHERE a.matricula=%s AND i.periodo=%s
    """, (matricula, periodo_activo))


    alumno = cursor.fetchone()

    cursor.execute("SELECT nombre FROM actividades")
    talleres = cursor.fetchall()

    conn.close()

    return render_template(
        "editar_inscripcion_admin.html",
        alumno=alumno,
        talleres=talleres,
        mensaje=""
    )

# ======================
# Borra datos
# ======================

@app.route('/limpiar_datos')
def limpiar_datos():

    if session.get('rol') != 'admin':
        return "Acceso denegado"

    conn = get_db()
    cursor = conn.cursor()

    # ✅ BORRAR EN ORDEN CORRECTO
    cursor.execute("DELETE FROM evaluaciones")
    cursor.execute("DELETE FROM resultados")
    cursor.execute("DELETE FROM inscripciones")

    # ⚠️ OPCIONAL (solo si quieres borrar alumnos)
    cursor.execute("DELETE FROM alumnos")

   

    conn.commit()
    conn.close()

    return "✅ Datos eliminados correctamente"



# ======================
# Importar estudiantes
# ======================

@app.route('/cargar_inscripciones_excel', methods=['GET','POST'])
def cargar_inscripciones_excel():

    if session.get('rol') != 'admin':
        return "Acceso denegado"

    mensaje = ""

    if request.method == 'POST':

        archivo = request.files.get('archivo')

        if not archivo or archivo.filename == '':
            return render_template("cargar_excel.html", mensaje="❌ Archivo no válido")

        try:
            df = pd.read_excel(archivo)

            # ✅ normalizar columnas
            df.columns = df.columns.str.strip().str.lower()

            # ✅ limpiar texto raro
            df = df.replace(r'&nbsp;|°', '', regex=True)

            conn = get_db()
            cursor = conn.cursor()

            insertados = 0
            duplicados = 0
            errores = 0

            for i, row in df.iterrows():

                try:
                    matricula = str(row.get('matricula', '')).strip()
                    nombre = str(row.get('nombre', '')).strip()
                    carrera = str(row.get('carrera', '')).strip()
                    semestre = str(row.get('semestre', '')).strip()
                    genero = str(row.get('genero', '')).strip()
                    telefono = str(row.get('telefono', '')).strip()
                    actividad = str(row.get('actividad', '')).strip()
                    periodo = str(row.get('periodo', '')).strip()

                    # ✅ validación básica
                    if not matricula or not actividad or not periodo:
                        errores += 1
                        continue

                    # ✅ insertar alumno si no existe
                    cursor.execute("SELECT 1 FROM alumnos WHERE matricula=%s", (matricula,))
                    if not cursor.fetchone():
                        cursor.execute("""
                            INSERT INTO alumnos (matricula, nombre, carrera)
                            VALUES (%s, %s, %s)
                        """, (matricula, nombre, carrera))

                    # ✅ insertar usuario con HASH 🔐
                    cursor.execute("SELECT 1 FROM usuarios WHERE usuario=%s", (matricula,))
                    if not cursor.fetchone():
                        password_hash = generate_password_hash(matricula)

                        cursor.execute("""
                            INSERT INTO usuarios (usuario, password, rol, cambio_password)
                            VALUES (%s, %s, 'alumno', 0)
                        """, (matricula, password_hash))

                    # ✅ validar duplicado de inscripción
                    cursor.execute("""
                        SELECT 1 FROM inscripciones
                        WHERE matricula=%s AND actividad=%s AND periodo=%s
                    """, (matricula, actividad, periodo))

                    if not cursor.fetchone():
                        cursor.execute("""
                            INSERT INTO inscripciones
                            (matricula, semestre, genero, telefono, actividad, periodo)
                            VALUES (%s,%s,%s,%s,%s,%s)
                        """, (matricula, semestre, genero, telefono, actividad, periodo))

                        insertados += 1
                    else:
                        duplicados += 1

                except Exception as e:
                    print(f"❌ Error fila {i}: {e}")
                    errores += 1

            conn.commit()
            conn.close()

            mensaje = f"""
✅ Insertados: {insertados}
⚠️ Duplicados: {duplicados}
❌ Errores: {errores}
"""

        except Exception as e:
            mensaje = f"❌ Error general: {str(e)}"

        return render_template("cargar_excel.html", mensaje=mensaje)

    return render_template("cargar_excel.html")

# ======================
# tODOS LOS FORMATOS
# ======================


@app.route('/formatos', methods=['GET', 'POST'])
def formatos():

    if 'usuario' not in session:
        return redirect('/')

    conn = get_db()
    cursor = conn.cursor()

    # ✅ periodo activo
    cursor.execute("SELECT periodo FROM periodos WHERE activo=1")
    dato = cursor.fetchone()
    periodo = dato[0] if dato else ""

    # ✅ obtener TODOS los talleres
    cursor.execute("""
    SELECT nombre FROM actividades
    WHERE docente=%s
    """, (session['usuario'],))

    talleres = [t[0] for t in cursor.fetchall()]
    taller_seleccionado = None

    # ✅ si solo tiene 1 → automático
    if len(talleres) == 1:
        taller_seleccionado = talleres[0]

    # ✅ si elige en formulario
    if request.method == 'POST':
        taller_seleccionado = request.form['taller']

    conn.close()

    return render_template(
        'formatos.html',
        talleres=talleres,
        taller_seleccionado=taller_seleccionado,
        periodo=periodo
    )

# ======================
# Exportar formato 5 multiple
# ======================


@app.route('/exportar_constancias_todos/<actividad>/<tipo>')
def exportar_constancias_todos(actividad, tipo):

    if 'usuario' not in session:
        return redirect('/')

    conn = get_db()
    cursor = conn.cursor()

    # ✅ periodo activo
    cursor.execute("SELECT periodo FROM periodos WHERE activo=1")
    dato = cursor.fetchone()
    periodo = dato[0] if dato else ""

    # ✅ CONSULTA CORREGIDA (JOIN + TRIM + LOWER ✅)
    cursor.execute("""
    SELECT alumnos.nombre,
       alumnos.matricula,
       alumnos.carrera,
       COALESCE(evaluaciones.promedio, 0),
       COALESCE(evaluaciones.nivel, ''),
       resultados.actividad
    FROM resultados
    JOIN alumnos ON alumnos.matricula = resultados.matricula
    LEFT JOIN evaluaciones 
    ON evaluaciones.matricula = resultados.matricula
    AND evaluaciones.actividad = resultados.actividad
    AND evaluaciones.periodo = resultados.periodo
    WHERE resultados.actividad=%s 
    AND resultados.periodo=%s
    AND TRIM(LOWER(resultados.resultado)) = 'acreditado'
    ORDER BY alumnos.nombre
    """, (actividad, periodo))


    datos = cursor.fetchall()
    conn.close()

    # ✅ validar
    if not datos:
        return "❌ No hay alumnos acreditados"

    from datetime import datetime
    fecha = datetime.now()

    # ✅ MESES EN ESPAÑOL
    meses = {
        "January": "enero",
        "February": "febrero",
        "March": "marzo",
        "April": "abril",
        "May": "mayo",
        "June": "junio",
        "July": "julio",
        "August": "agosto",
        "September": "septiembre",
        "October": "octubre",
        "November": "noviembre",
        "December": "diciembre"
    }

    mes = meses[fecha.strftime("%B")]

    lista = []

    for i, d in enumerate(datos, start=1):
        lista.append({
            "numero": f"{i:03}",
            "número": f"{i:03}",
            "nombre_estudiante": d[0],
            "estudiante": d[0],
            "matricula": d[1],
            "carrera": d[2],
            "nivel": d[4],
            "promedio": round(d[3], 2),
            "periodo": periodo,
            "día_número": fecha.day,
            "mes": mes,
            "año": fecha.year,
            "TALLER": d[5]
        })

    from docxtpl import DocxTemplate
    from flask import send_file

    # ✅ seleccionar formato
    if tipo == "01":
        doc = DocxTemplate("Formato 05-01.docx")
    else:
        doc = DocxTemplate("Formato 05-02.docx")

    contexto = {
        "alumnos": lista
    }

    doc.render(contexto)

    archivo = f"constancias_{actividad}.docx"
    doc.save(archivo)

    return send_file(archivo, as_attachment=True)

    

# ======================
# Exportar formato 4 multiple
# ======================


@app.route('/exportar_todos_formato4/<actividad>')
def exportar_todos_formato4(actividad):

    if 'usuario' not in session:
        return redirect('/')

    conn = get_db()
    cursor = conn.cursor()

    # ✅ periodo activo
    cursor.execute("SELECT periodo FROM periodos WHERE activo=1")
    dato = cursor.fetchone()
    periodo = dato[0] if dato else ""

    # ✅ CONSULTA CORRECTA (SIN SELECT *)
    cursor.execute("""
    SELECT alumnos.nombre,
           evaluaciones.c1,
           evaluaciones.c2,
           evaluaciones.c3,
           evaluaciones.c4,
           evaluaciones.c5,
           evaluaciones.c6,
           evaluaciones.c7,
           evaluaciones.promedio,
           evaluaciones.nivel,
           evaluaciones.observaciones
    FROM evaluaciones
    INNER JOIN alumnos ON alumnos.matricula = evaluaciones.matricula
    WHERE evaluaciones.actividad=%s AND evaluaciones.periodo=%s
    """, (actividad, periodo))

    datos = cursor.fetchall()
    conn.close()

    print("DATOS:", datos)  # ✅ DEBUG

    if not datos:
        return "❌ No hay evaluaciones registradas"

    def marca(valor, esperado):
        return "X" if valor == esperado else ""

    lista = []

    for d in datos:
        lista.append({
            "nombre": d[0],
            "actividad": actividad,
            "periodo": periodo,

            "c1_e": marca(d[1],0), "c1_n": marca(d[1],1), "c1_b": marca(d[1],2), "c1_s": marca(d[1],3), "c1_i": marca(d[1],4),
            "c2_e": marca(d[2],0), "c2_n": marca(d[2],1), "c2_b": marca(d[2],2), "c2_s": marca(d[2],3), "c2_i": marca(d[2],4),
            "c3_e": marca(d[3],0), "c3_n": marca(d[3],1), "c3_b": marca(d[3],2), "c3_s": marca(d[3],3), "c3_i": marca(d[3],4),
            "c4_e": marca(d[4],0), "c4_n": marca(d[4],1), "c4_b": marca(d[4],2), "c4_s": marca(d[4],3), "c4_i": marca(d[4],4),
            "c5_e": marca(d[5],0), "c5_n": marca(d[5],1), "c5_b": marca(d[5],2), "c5_s": marca(d[5],3), "c5_i": marca(d[5],4),
            "c6_e": marca(d[6],0), "c6_n": marca(d[6],1), "c6_b": marca(d[6],2), "c6_s": marca(d[6],3), "c6_i": marca(d[6],4),
            "c7_e": marca(d[7],0), "c7_n": marca(d[7],1), "c7_b": marca(d[7],2), "c7_s": marca(d[7],3), "c7_i": marca(d[7],4),

            "promedio": round(d[8],2),
            "nivel": d[9],
            "observaciones": d[10]
        })

    print("LISTA:", lista)  # ✅ DEBUG

    from docxtpl import DocxTemplate
    from flask import send_file

    doc = DocxTemplate("formato_4_multiple.docx")

    contexto = {
        "alumnos": lista
    }

    doc.render(contexto)

    archivo = "formato4_todos.docx"
    doc.save(archivo)

    return send_file(archivo, as_attachment=True)

# ======================
# Eliminar evaluación
# ======================

@app.route('/eliminar_evaluacion/<matricula>/<actividad>')
def eliminar_evaluacion(matricula, actividad):

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT periodo FROM periodos WHERE activo=1")
    periodo = cursor.fetchone()[0]

    cursor.execute("""
    DELETE FROM evaluaciones
    WHERE matricula=%s AND actividad=%s AND periodo=%s
    """, (matricula, actividad, periodo))

    conn.commit()
    conn.close()

    return redirect(f"/acreditacion?taller={actividad}")

# ======================
# formulario evaluación
# ======================


@app.route('/evaluacion/<int:index>')
def evaluacion(index):

    if 'usuario' not in session:
        return redirect('/')

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT periodo FROM periodos WHERE activo=1")
    periodo = cursor.fetchone()[0]

    docente = session['usuario']

    cursor.execute("""
    SELECT alumnos.matricula, alumnos.nombre, inscripciones.actividad
    FROM inscripciones
    JOIN alumnos ON alumnos.matricula = inscripciones.matricula
    WHERE inscripciones.actividad IN (
        SELECT nombre FROM actividades WHERE docente=%s
    )
    AND inscripciones.periodo=%s
    """, (docente, periodo))

    lista = cursor.fetchall()

    if index >= len(lista):
        conn.close()
        return redirect('/acreditacion')

    alumno = lista[index]

    # ✅ VERIFICAR SI YA EXISTE EVALUACIÓN
    cursor.execute("""
    SELECT id FROM evaluaciones
    WHERE matricula=%s AND actividad=%s AND periodo=%s
    """, (alumno[0], alumno[2], periodo))

    existe = cursor.fetchone()

    conn.close()

    if existe:
        return "❌ Este alumno ya fue evaluado"

    return render_template(
        "evaluacion.html",
        alumno=alumno,
        index=index,
        total=len(lista)
    )


# ======================
# Guardar evaluación
# ======================

@app.route('/guardar_evaluacion', methods=['POST'])
def guardar_evaluacion():

    conn = get_db()
    cursor = conn.cursor()

    # ✅ periodo activo
    cursor.execute("SELECT periodo FROM periodos WHERE activo=1")
    periodo = cursor.fetchone()[0]

    index = int(request.form['index'])

    c1 = int(request.form['c1'])
    c2 = int(request.form['c2'])
    c3 = int(request.form['c3'])
    c4 = int(request.form['c4'])
    c5 = int(request.form['c5'])
    c6 = int(request.form['c6'])
    c7 = int(request.form['c7'])

    print(c1, c2, c3, c4, c5, c6, c7)

    promedio = (c1+c2+c3+c4+c5+c6+c7)/7

    nivel = calcular_nivel(promedio)

    cursor.execute("""
    INSERT INTO evaluaciones
    (matricula, actividad, periodo, c1,c2,c3,c4,c5,c6,c7, promedio, nivel, observaciones)
    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """,(
        request.form['matricula'],
        request.form['actividad'],
        periodo,
        c1,c2,c3,c4,c5,c6,c7,
        promedio,
        nivel,
        request.form['observaciones']
    ))

    conn.commit()
    conn.close()

    # ✅ IR AL SIGUIENTE
    return redirect(f"/evaluacion/{index+1}")



# ======================
# Calcular promedio
# ======================
def calcular_nivel(promedio):

    if promedio >= 3.5:
        return "Excelente"
    elif promedio >= 2.5:
        return "Notable"
    elif promedio >= 1.5:
        return "Bueno"
    elif promedio >= 1.0:
        return "Suficiente"
    else:
        return "Insuficiente"


# ======================
# gUARDAR RESULTADO
# ======================

@app.route('/guardar_resultado', methods=['POST'])
def guardar_resultado():

    if 'usuario' not in session:
        return redirect('/')

    conn = get_db()
    cursor = conn.cursor()

    matricula = request.form['matricula']
    actividad = request.form['actividad']
    resultado = request.form['resultado']

    # ✅ SIEMPRE usar periodo activo (NO el del HTML)
    cursor.execute("SELECT periodo FROM periodos WHERE activo=1")
    periodo = cursor.fetchone()[0]

    # ✅ verificar si ya existe
    cursor.execute("""
    SELECT id FROM resultados
    WHERE matricula=%s AND actividad=%s AND periodo=%s
    """, (matricula, actividad, periodo))

    existe = cursor.fetchone()

    if existe:
        # ✅ actualizar
        cursor.execute("""
        UPDATE resultados
        SET resultado=%s
        WHERE matricula=%s AND actividad=%s AND periodo=%s
        """, (resultado, matricula, actividad, periodo))
    else:
        # ✅ insertar
        cursor.execute("""
        INSERT INTO resultados (matricula, actividad, periodo, resultado)
        VALUES (%s,%s,%s,%s)
        """, (matricula, actividad, periodo, resultado))

    conn.commit()
    conn.close()
    
    return redirect(f"/acreditacion?taller={actividad}&ok=1")


# ======================
# Formato 4
# ======================



@app.route('/exportar_formato4/<matricula>/<actividad>')
def exportar_formato4(matricula, actividad):

    if 'usuario' not in session:
        return redirect('/')

    conn = get_db()
    cursor = conn.cursor()

    # ✅ periodo activo
    cursor.execute("SELECT periodo FROM periodos WHERE activo=1")
    periodo = cursor.fetchone()[0]

    # ✅ datos alumno
    cursor.execute("""
    SELECT nombre FROM alumnos
    WHERE matricula=%s
    """, (matricula,))
    alumno = cursor.fetchone()[0]

    # ✅ obtener evaluación
    cursor.execute("""
    SELECT c1,c2,c3,c4,c5,c6,c7,promedio,nivel,observaciones
    FROM evaluaciones
    WHERE matricula=%s AND actividad=%s AND periodo=%s
    """, (matricula, actividad, periodo))

    data = cursor.fetchone()
    conn.close()

    if not data:
        return "❌ No hay evaluación registrada"

    c1,c2,c3,c4,c5,c6,c7,promedio,nivel,obs = data

    # ✅ función para marcar X
    def marca(valor, esperado):
        return "X" if valor == esperado else ""

    from docxtpl import DocxTemplate
    doc = DocxTemplate("formato_4.docx")

    contexto = {
        "nombre": alumno,
        "actividad": actividad,
        "periodo": periodo,

        "c1_e": marca(c1,0), "c1_n": marca(c1,1), "c1_b": marca(c1,2), "c1_s": marca(c1,3), "c1_i": marca(c1,4),
        "c2_e": marca(c2,0), "c2_n": marca(c2,1), "c2_b": marca(c2,2), "c2_s": marca(c2,3), "c2_i": marca(c2,4),
        "c3_e": marca(c3,0), "c3_n": marca(c3,1), "c3_b": marca(c3,2), "c3_s": marca(c3,3), "c3_i": marca(c3,4),
        "c4_e": marca(c4,0), "c4_n": marca(c4,1), "c4_b": marca(c4,2), "c4_s": marca(c4,3), "c4_i": marca(c4,4),
        "c5_e": marca(c5,0), "c5_n": marca(c5,1), "c5_b": marca(c5,2), "c5_s": marca(c5,3), "c5_i": marca(c5,4),
        "c6_e": marca(c6,0), "c6_n": marca(c6,1), "c6_b": marca(c6,2), "c6_s": marca(c6,3), "c6_i": marca(c6,4),
        "c7_e": marca(c7,0), "c7_n": marca(c7,1), "c7_b": marca(c7,2), "c7_s": marca(c7,3), "c7_i": marca(c7,4),

        "promedio": round(promedio,2),
        "nivel": nivel,
        "observaciones": obs
    }

    doc.render(contexto)

    archivo = f"formato4_{matricula}.docx"
    doc.save(archivo)

    return send_file(archivo, as_attachment=True)



# ======================
# Tabla editar evento docente
# ======================
@app.route('/editar_evento/<int:id>', methods=['GET','POST'])
def editar_evento(id):

    if 'usuario' not in session:
        return redirect('/')

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'POST':

        cursor.execute("""
        UPDATE eventos SET
        nombre_evento=%s,
        institucion=%s,
        fecha=%s,
        participantes=%s,
        mujeres=%s,
        hombres=%s,
        resultados=%s
        WHERE id=%s
        """, (
            request.form['nombre_evento'],
            request.form['institucion'],
            request.form['fecha'],
            request.form['participantes'],
            request.form['mujeres'],
            request.form['hombres'],
            request.form['resultados'],
            id
        ))

        conn.commit()
        conn.close()

        return redirect('/eventos_docente')

    # ✅ traer datos actuales
    cursor.execute("SELECT * FROM eventos WHERE id=%s", (id,))
    evento = cursor.fetchone()

    conn.close()

    return render_template('editar_evento.html', evento=evento)

# ======================
# fORMATO 3
# ======================


@app.route('/exportar_resultados/<actividad>')
def exportar_resultados(actividad):

    if 'usuario' not in session:
        return redirect('/')

    conn = get_db()
    cursor = conn.cursor()

    # ✅ obtener periodo activo
    cursor.execute("SELECT periodo FROM periodos WHERE activo=1")
    periodo = cursor.fetchone()[0]

    # ✅ CONSULTA CORREGIDA ✅
    cursor.execute("""
    SELECT alumnos.nombre, alumnos.matricula,
           alumnos.carrera, inscripciones.semestre,
           COALESCE(MAX(resultados.resultado),'No Acreditado') AS resultado
    FROM inscripciones
    LEFT JOIN resultados 
        ON resultados.matricula = inscripciones.matricula
        AND resultados.actividad = inscripciones.actividad
        AND resultados.periodo = inscripciones.periodo
    JOIN alumnos ON alumnos.matricula = inscripciones.matricula
    WHERE inscripciones.actividad=%s 
      AND inscripciones.periodo=%s
    GROUP BY alumnos.nombre, alumnos.matricula, alumnos.carrera, inscripciones.semestre
    ORDER BY alumnos.nombre
    """, (actividad, periodo))

    datos = cursor.fetchall()
    conn.close()

    # ✅ validar
    if not datos:
        return "❌ No hay datos para generar el formato"

    from docxtpl import DocxTemplate
    from datetime import datetime
    from flask import send_file

    doc = DocxTemplate("formato_resultados.docx")

    tabla = []

    for i, d in enumerate(datos, start=1):
        tabla.append({
            "no": i,
            "nombre": d[0],
            "control": d[1],
            "carrera": d[2],
            "semestre": d[3],
            "resultado": d[4]
        })

    contexto = {
        "tabla": tabla,
        "actividad": actividad,
        "fecha": datetime.now().strftime("%d/%m/%Y")
    }

    doc.render(contexto)

    archivo = "resultados.docx"
    doc.save(archivo)

    return send_file(archivo, as_attachment=True)


# ======================
# ACEDRITACION FORMATO 3
# ======================

@app.route('/acreditacion')
def acreditacion():

    if 'usuario' not in session:
        return redirect('/')

    if session.get('rol') != 'docente':
        return "Acceso denegado"

    conn = get_db()
    cursor = conn.cursor()

    docente = session['usuario']

    # ✅ obtener periodo activo
    cursor.execute("SELECT periodo FROM periodos WHERE activo=1")
    dato = cursor.fetchone()
    periodo = dato[0] if dato else ""

    # ✅ obtener taller desde URL
    taller = request.args.get('taller')

    if not taller:
        conn.close()
        return "❌ Debes seleccionar un taller"

    # ✅ CONSULTA FILTRADA POR TALLER (CLAVE)
    cursor.execute("""
    SELECT DISTINCT 
        alumnos.matricula, 
        alumnos.nombre, 
        alumnos.carrera,
        inscripciones.semestre, 
        inscripciones.actividad,
        COALESCE(resultados.resultado, 'Sin evaluar')
    FROM inscripciones
    JOIN alumnos ON alumnos.matricula = inscripciones.matricula
    LEFT JOIN resultados 
        ON resultados.matricula = inscripciones.matricula
        AND resultados.actividad = inscripciones.actividad
        AND resultados.periodo = inscripciones.periodo
    WHERE inscripciones.actividad = %s
    AND inscripciones.periodo = %s
    """, (taller, periodo))

    alumnos = cursor.fetchall()

    conn.close()

    return render_template(
        'acreditacion.html',
        alumnos=alumnos,
        taller=taller
    )

# ======================
# Tabla evento docente
# ======================

@app.route('/eventos_docente')
def eventos_docente():

    if 'usuario' not in session:
        return redirect('/')

    if session.get('rol') != 'docente':
        return "Acceso denegado"

    conn = get_db()
    cursor = conn.cursor()

    
    cursor.execute("SELECT periodo FROM periodos WHERE activo=1")
    periodo = cursor.fetchone()[0]

    cursor.execute("""
    SELECT id, nombre_evento, institucion, fecha,
       participantes, mujeres, hombres, resultados
    FROM eventos
    WHERE docente=%s AND periodo=%s
    ORDER BY fecha DESC
    """, (session['usuario'], periodo))


    eventos = cursor.fetchall()

    conn.close()

    return render_template('eventos_docente.html',
                           eventos=eventos)

# ======================
# Word evento
# ======================

@app.route('/registrar_evento', methods=['POST'])
def registrar_evento():

    if 'usuario' not in session:
        return redirect('/')

    conn = get_db()
    cursor = conn.cursor()

    # ✅ obtener periodo activo
    cursor.execute("SELECT periodo FROM periodos WHERE activo=1")
    dato = cursor.fetchone()
    periodo = dato[0] if dato else ""

    # ✅ OBTENER ACTIVIDAD DESDE FORM (IMPORTANTE)
    actividad = request.form.get('actividad', 'Taller X')

    cursor.execute("""
    INSERT INTO eventos 
    (docente, actividad, nombre_evento, institucion, fecha,
     participantes, mujeres, hombres, resultados, periodo)
    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        session['usuario'],
        actividad,   # ✅ ya no fijo
        request.form['nombre_evento'],
        request.form['institucion'],
        request.form['fecha'],
        request.form['participantes'],
        request.form['mujeres'],
        request.form['hombres'],
        request.form['resultados'],
        periodo   # ✅ periodo dinámico
    ))

    conn.commit()
    conn.close()

    return redirect('/docente')

# ======================
# Word evento registar
# ======================

@app.route('/registrar_evento', methods=['GET'])
def mostrar_registro_evento():

    if 'usuario' not in session:
        return redirect('/')

    if session.get('rol') != 'docente':
        return "Acceso denegado"

    return render_template('registrar_evento.html')



from docxtpl import DocxTemplate
from flask import send_file, redirect, session
from datetime import datetime


# ======================
# Exportar Informe
# ======================



@app.route('/exportar_informe/<actividad>')
def exportar_informe(actividad):

    if 'usuario' not in session:
        return redirect('/')

    conn = get_db()
    cursor = conn.cursor()

    # ✅ PERIODO ACTIVO
    cursor.execute("SELECT periodo FROM periodos WHERE activo=1")
    periodo = cursor.fetchone()[0]

    # ✅ CONSULTA ROBUSTA (SIN FALLAS)
    cursor.execute("""
    SELECT nombre_evento, institucion, fecha,
           participantes, mujeres, hombres, resultados, actividad, periodo
    FROM eventos
    WHERE REPLACE(LOWER(actividad),'á','a') = REPLACE(LOWER(%s),'á','a')
    """, (actividad,))

    datos = cursor.fetchall()

    if not datos:
        conn.close()
        return "❌ No hay eventos registrados para este taller"

    conn.close()

    # ✅ FILTRAR EN PYTHON POR PERIODO (más seguro)
    datos_filtrados = [d for d in datos if d[8] == periodo]

    # 🔴 si no hay en ese periodo, usa todos (opcional)
    if not datos_filtrados:
        datos_filtrados = datos

    # ✅ detectar tipo taller
    actividad_nombre = actividad.upper()

    if "FUT" in actividad_nombre or "VOLE" in actividad_nombre or "BASQU" in actividad_nombre:
        cultural = ""
        deportiva = "X"
    else:
        cultural = "X"
        deportiva = ""

    from docxtpl import DocxTemplate
    from datetime import datetime
    from flask import send_file

    doc = DocxTemplate("formato_informe.docx")

    tabla = []

    for i, d in enumerate(datos_filtrados, start=1):
        tabla.append({
            "no": i,
            "evento": d[0],
            "institucion": d[1],
            "fecha_evento": d[2],
            "participantes": d[3],
            "mujeres": d[4],
            "hombres": d[5],
            "resultados": d[6]
        })

    contexto = {
        "tabla": tabla,
        "periodo": periodo,
        "fecha": datetime.now().strftime("%d/%m/%Y"),
        "docente": session['usuario'],
        "actividad": actividad,
        "cultural": cultural,
        "deportiva": deportiva
    }

    doc.render(contexto)

    archivo = f"informe_{actividad}.docx"
    doc.save(archivo)

    return send_file(archivo, as_attachment=True)




# ======================
# Word lista
# ======================

from docxtpl import DocxTemplate
from flask import send_file
from datetime import datetime

@app.route('/exportar_word/<actividad>')
def exportar_word(actividad):

    if 'usuario' not in session:
        return redirect('/')

    conn = get_db()
    cursor = conn.cursor()

    
    cursor.execute("SELECT periodo FROM periodos WHERE activo=1")
    periodo = cursor.fetchone()[0]

    cursor.execute("""
    SELECT alumnos.nombre, alumnos.matricula,
       alumnos.carrera, inscripciones.semestre
    FROM inscripciones
    JOIN alumnos ON alumnos.matricula = inscripciones.matricula
    WHERE inscripciones.actividad=%s AND inscripciones.periodo=%s
    """, (actividad, periodo))


    datos = cursor.fetchall()
    conn.close()

    
    if not datos:
        return "❌ No hay alumnos registrados en este taller"


    # ✅ Cargar plantilla
    doc = DocxTemplate("formato_base.docx")

    tabla = []

    for i, fila in enumerate(datos, start=1):
        tabla.append({
            "no": i,
            "nombre": fila[0],
            "control": fila[1],
            "carrera": fila[2],
            "semestre": fila[3],
            "obs": ""
        })

    
    contexto = {
    "actividad": actividad,
    "fecha": datetime.now().strftime("%d/%m/%Y"),
    "docente": session['usuario'],
    "tabla": tabla
    }


    doc.render(contexto)

    archivo = f"formato_{actividad}.docx"
    doc.save(archivo)

    return send_file(archivo, as_attachment=True)

# ======================
# fechas 
# ======================

def obtener_estado_fechas():

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT fecha_inicio, fecha_fin FROM control_fechas LIMIT 1")
    data = cursor.fetchone()

    conn.close()

    if not data:
        return "sin_configurar"

    inicio = datetime.strptime(data[0], "%Y-%m-%d")
    fin = datetime.strptime(data[1], "%Y-%m-%d")
    hoy = datetime.now()

    if hoy < inicio:
        return "no_iniciado"
    elif hoy > fin:
        return "cerrado"
    else:
        return "activo"


# ======================
# contador de dias restantes
# ======================

from datetime import datetime

def obtener_tiempo_restante():

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT fecha_inicio, fecha_fin FROM control_fechas LIMIT 1")
    data = cursor.fetchone()

    conn.close()

    if not data:
        return "No configurado", "sin_configurar"

    inicio = datetime.strptime(data[0], "%Y-%m-%d")
    fin = datetime.strptime(data[1], "%Y-%m-%d")
    hoy = datetime.now()

    if hoy < inicio:
        restante = inicio - hoy
        estado = "no_iniciado"
    elif hoy > fin:
        restante = fin - hoy
        estado = "cerrado"
    else:
        restante = fin - hoy
        estado = "activo"

    # ✅ FORMATEAR TIEMPO
    if restante.total_seconds() < 0:
        restante_str = "0 días"
    else:
        dias = restante.days
        horas = restante.seconds // 3600
        minutos = (restante.seconds % 3600) // 60

        restante_str = f"{dias} días, {horas} horas, {minutos} minutos"

    return restante_str, estado

# ======================
# editar fechas
# ======================
@app.route('/editar_fechas', methods=['GET','POST'])
def editar_fechas():

    if session.get('rol') != 'admin':
        return "Acceso denegado"

    conn = get_db()
    cursor = conn.cursor()

    mensaje = ""
    data = None

    if request.method == 'POST':

        inicio = request.form['inicio']
        fin = request.form['fin']

        if not inicio or not fin:
            mensaje = "❌ Debes llenar ambas fechas"

        elif inicio > fin:
            mensaje = "❌ La fecha de inicio no puede ser mayor que la de cierre"

        else:
            cursor.execute("DELETE FROM control_fechas")

            cursor.execute("""
            INSERT INTO control_fechas (fecha_inicio, fecha_fin)
            VALUES (%s, %s)
            """, (inicio, fin))

            conn.commit()

            mensaje = "✅ Fechas actualizadas correctamente"

    # ✅ SIEMPRE CONSULTAR PARA MOSTRAR
    cursor.execute("SELECT fecha_inicio, fecha_fin FROM control_fechas LIMIT 1")
    data = cursor.fetchone()

    conn.close()

    # ✅ SIEMPRE DEVOLVER ALGO
    return render_template("editar_fechas.html", datos=data, mensaje=mensaje)




# ======================
# DATOS BASE
# ======================
def cargar_datos():
    conn = get_db()
    cursor = conn.cursor()

    password_hash = generate_password_hash('DJJazz1994')

    cursor.execute("""
    INSERT INTO usuarios (usuario, password, rol, cambio_password)
    VALUES (%s, %s, %s, %s)
    ON CONFLICT (usuario)
    DO UPDATE SET 
        password = EXCLUDED.password,
        rol = EXCLUDED.rol,
        cambio_password = EXCLUDED.cambio_password
    """, ('admin', password_hash, 'admin', 1))

    conn.commit()
    conn.close()


# ======================
# LOGIN
# ======================
from werkzeug.security import check_password_hash

@app.route('/', methods=['GET','POST'])
def login():

    if request.method == 'POST':
        usuario = request.form['usuario']
        password = request.form['password']

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("""
        SELECT password, rol, cambio_password 
        FROM usuarios 
        WHERE usuario=%s
        """, (usuario,))

        user = cursor.fetchone()
        conn.close()  # ✅ cerrar antes

        # ✅ validar usuario y password
        if user and check_password_hash(user[0], password):

            session['usuario'] = usuario
            session['rol'] = user[1]

            if user[2] == 0:
                return redirect('/cambiar_password')

            return redirect('/' + user[1])

        # ❌ login incorrecto
        return "❌ Usuario o contraseña incorrectos"

    # ✅ GET
    restante, estado = obtener_tiempo_restante()

    return render_template(
        'login.html',
        estado=estado,
        restante=restante
    )


# ======================
# Manejo de fechas de apertura y cierre 
# ======================

@app.route('/fecha_apertura', methods=['GET','POST'])
def fecha_apertura():

    if session.get('rol') != 'admin':
        return "Acceso denegado"

    mensaje = ""

    if request.method == 'POST':

        inicio = request.form['inicio']
        fin = request.form['fin']

        if not inicio or not fin:
            mensaje = "❌ Debes llenar ambas fechas"
            return render_template("fecha_apertura.html", mensaje=mensaje)

        # 🔴 VALIDACIÓN: inicio menor que fin
        if inicio > fin:
            mensaje = "❌ La fecha de inicio no puede ser mayor que la de cierre"
            return render_template("fecha_apertura.html", mensaje=mensaje)

        conn = get_db()
        cursor = conn.cursor()

        # ✅ OPCIÓN PROFESIONAL: SOLO UNA CONFIGURACIÓN
        cursor.execute("DELETE FROM control_fechas")

        cursor.execute("""
        INSERT INTO control_fechas (fecha_inicio, fecha_fin)
        VALUES (%s, %s)
        """, (inicio, fin))

        conn.commit()
        conn.close()

        mensaje = "✅ Fechas guardadas correctamente"

    return render_template("fecha_apertura.html", mensaje=mensaje)


# ======================
# CAMBIO PASSWORD
# ======================
@app.route('/cambiar_password', methods=['GET','POST'])
def cambiar_password():

    if request.method == 'POST':
        nueva = request.form['nuevo']

        conn = get_db()
        cursor = conn.cursor()

        nueva_hash = generate_password_hash(nueva)

        cursor.execute("""
        UPDATE usuarios SET password=%s, cambio_password=1 
        WHERE usuario=%s
        """, (nueva_hash, session['usuario']))


        conn.commit()
        conn.close()

        return redirect('/')

    return render_template('cambiar_password.html')

# ======================
# PERFIL ADMIN
# ======================
@app.route('/perfil_admin', methods=['GET','POST'])
def perfil_admin():

    if session.get('rol') != 'admin':
        return "Acceso denegado"

    mensaje = ""

    if request.method == 'POST':

        nuevo_usuario = request.form.get('usuario', '').strip()
        nueva_password = request.form.get('password', '').strip()

        # ✅ validaciones
        if not nuevo_usuario or not nueva_password:
            mensaje = "❌ Debes llenar todos los campos"
            return render_template("perfil_admin.html",
                                   usuario=session['usuario'],
                                   mensaje=mensaje)

        conn = get_db()
        cursor = conn.cursor()

        try:
            # 🔍 verificar si usuario ya existe (y no es el mismo)
            cursor.execute("""
                SELECT 1 FROM usuarios 
                WHERE usuario=%s AND usuario!=%s
            """, (nuevo_usuario, session['usuario']))

            if cursor.fetchone():
                conn.close()
                mensaje = "❌ El usuario ya existe"
                return render_template("perfil_admin.html",
                                       usuario=session['usuario'],
                                       mensaje=mensaje)

            # ✅ hash de password
            password_hash = generate_password_hash(nueva_password)

            # ✅ actualizar
            cursor.execute("""
                UPDATE usuarios 
                SET usuario=%s, password=%s 
                WHERE usuario=%s
            """, (nuevo_usuario, password_hash, session['usuario']))

            conn.commit()
            conn.close()

            # ✅ actualizar sesión
            session['usuario'] = nuevo_usuario

            return redirect('/admin')

        except Exception as e:
            conn.close()
            return f"❌ Error: {str(e)}"

    return render_template("perfil_admin.html", usuario=session['usuario'])

# ======================
# ADMIN
# ======================
@app.route('/admin')
def admin():

    if 'usuario' not in session:
        return redirect('/')

    if session.get('rol') != 'admin':
        return "Acceso denegado"

    return render_template('admin.html')

# ======================
# Docente
# ======================

@app.route('/docente', methods=['GET','POST'])
def docente():

    if 'usuario' not in session:
        return redirect('/')

    if session.get('rol') != 'docente':
        return "Acceso denegado"

    conn = get_db()
    cursor = conn.cursor()

    docente = session['usuario']

    # ✅ obtener periodo activo (CORREGIDO)
    cursor.execute("SELECT periodo FROM periodos WHERE activo=1")
    dato = cursor.fetchone()
    periodo_seleccionado = dato[0] if dato else "No definido"

    # ✅ obtener talleres del docente
    cursor.execute("""
    SELECT nombre FROM actividades
    WHERE docente=%s
    """, (docente,))

    talleres = [t[0] for t in cursor.fetchall()]

    alumnos_lista = []
    taller_seleccionado = None

    # ✅ si solo tiene 1 taller → automático
    if len(talleres) == 1:
        taller_seleccionado = talleres[0]

    # ✅ si selecciona desde formulario
    if request.method == 'POST':
        taller_seleccionado = request.form['taller']

    # ✅ si ya hay taller seleccionado → cargar alumnos SOLO de ese
    if taller_seleccionado and periodo_seleccionado != "No definido":
        cursor.execute("""
        SELECT alumnos.matricula, alumnos.nombre, alumnos.carrera,
               inscripciones.genero, inscripciones.telefono
        FROM inscripciones
        JOIN alumnos ON alumnos.matricula = inscripciones.matricula
        WHERE inscripciones.actividad=%s AND inscripciones.periodo=%s
        """, (taller_seleccionado, periodo_seleccionado))

        alumnos_lista = cursor.fetchall()

    conn.close()

    return render_template('docente.html',
        usuario=docente,
        talleres=talleres,
        alumnos=alumnos_lista,
        taller_seleccionado=taller_seleccionado,
        periodo_seleccionado=periodo_seleccionado   
    )
# ======================
# crear Docente
# ======================

@app.route('/crear_docente', methods=['GET','POST'])
def crear_docente():

    if 'usuario' not in session:
        return redirect('/')

    if session.get('rol') != 'admin':
        return "Acceso denegado"

    mensaje = ""

    if request.method == 'POST':

        usuario = request.form['usuario']
        password = request.form['password']

        if not usuario or not password:
            mensaje = "❌ Debes llenar todos los campos"
            return render_template("crear_docente.html", mensaje=mensaje)

        conn = get_db()
        cursor = conn.cursor()

        # ✅ verificar si ya existe
        cursor.execute("SELECT * FROM usuarios WHERE usuario=%s", (usuario,))
        if cursor.fetchone():
            conn.close()
            mensaje = "❌ El usuario ya existe"
            return render_template("crear_docente.html", mensaje=mensaje)

        # ✅ insertar docente
        password_hash = generate_password_hash(password)

        cursor.execute("""
        INSERT INTO usuarios (usuario, password, rol, cambio_password)
        VALUES (%s, %s, 'docente', 0)
        """, (usuario, password_hash))

        conn.commit()
        conn.close()

        mensaje = "✅ Docente creado correctamente"

    return render_template("crear_docente.html", mensaje=mensaje)

# ======================
# ===== Ver docente =====
# ======================


@app.route('/ver_docentes')
def ver_docentes():

    if 'usuario' not in session:
        return redirect('/')

    if session.get('rol') != 'admin':
        return "Acceso denegado"

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT usuario FROM usuarios WHERE rol='docente'")
    docentes = cursor.fetchall()

    conn.close()

    return render_template('ver_docentes.html', docentes=docentes)

# ======================
# ===== Editar docente =====
# ======================

@app.route('/editar_docente/<usuario>', methods=['GET','POST'])
def editar_docente(usuario):

    if session.get('rol') != 'admin':
        return "Acceso denegado"

    mensaje = ""

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'POST':

        nueva_password = request.form.get('password', '').strip()

        # ✅ validar
        if not nueva_password:
            conn.close()
            mensaje = "❌ Debes ingresar una contraseña"
            return render_template('editar_docente.html', usuario=usuario, mensaje=mensaje)

        try:
            # ✅ generar hash
            password_hash = generate_password_hash(nueva_password)

            # ✅ actualizar password
            cursor.execute("""
            UPDATE usuarios 
            SET password=%s 
            WHERE usuario=%s
            """, (password_hash, usuario))

            conn.commit()
            conn.close()

            return redirect('/ver_docentes')

        except Exception as e:
            conn.close()
            return f"❌ Error: {str(e)}"

    conn.close()
    return render_template('editar_docente.html', usuario=usuario)



# ======================
# ===== eLIMNAR DOCENTE =====
# ======================


@app.route('/eliminar_docente/<usuario>')
def eliminar_docente(usuario):

    if 'usuario' not in session:
        return redirect('/')

    if session.get('rol') != 'admin':
        return "Acceso denegado"

    conn = get_db()
    cursor = conn.cursor()

    # ✅ Quitar docente de talleres antes de borrar
    cursor.execute("""
    UPDATE actividades SET docente=NULL WHERE docente=%s
    """, (usuario,))

    # ✅ Eliminar usuario docente
    cursor.execute("""
    DELETE FROM usuarios WHERE usuario=%s AND rol='docente'
    """, (usuario,))

    conn.commit()
    conn.close()

    return redirect('/ver_docentes')

# ======================
# ===== DOCENTE ASIGANAR TALLER  =====
# ======================
@app.route('/asignar_taller/<int:id>', methods=['GET','POST'])
def asignar_taller(id):

    if 'usuario' not in session:
        return redirect('/')

    if session.get('rol') != 'admin':
        return "Acceso denegado"

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'POST':

        docente = request.form['docente']

        cursor.execute("""
        UPDATE actividades SET docente=%s WHERE id=%s
        """, (docente, id))

        conn.commit()
        conn.close()

        return redirect('/admin_talleres')

    cursor.execute("SELECT usuario FROM usuarios WHERE rol='docente'")
    docentes = cursor.fetchall()

    conn.close()

    return render_template('asignar_taller.html', docentes=docentes, id=id)


# ======================
# ===== ALUMNOS =====
# ======================
@app.route('/admin_alumnos', methods=['GET','POST'])
def admin_alumnos():

    conn = get_db()
    cursor = conn.cursor()

    alumnos = []

    if request.method == 'POST':
        busqueda = request.form['busqueda']

        cursor.execute("""
        SELECT matricula, nombre, carrera 
        FROM alumnos
        WHERE matricula LIKE %s OR nombre LIKE %s
        """, ('%' + busqueda + '%', '%' + busqueda + '%'))

        alumnos = cursor.fetchall()

    conn.close()

    return render_template('admin_alumnos.html', alumnos=alumnos)

# ======================
# EDITAR ALUMNO
# ======================
@app.route('/editar_alumno/<matricula>', methods=['GET','POST'])
def editar_alumno(matricula):

    conn = get_db()
    cursor = conn.cursor()

    mensaje = ""

    if request.method == 'POST':

        nombre = request.form.get('nombre', '').strip()
        carrera = request.form.get('carrera', '').strip()
        nueva_password = request.form.get('password', '').strip()

        # ✅ validación
        if not nombre or not carrera:
            conn.close()
            mensaje = "❌ Nombre y carrera son obligatorios"
            return render_template('editar_alumno.html', alumno=(matricula, nombre, carrera), mensaje=mensaje)

        try:
            # ✅ actualizar datos alumno
            cursor.execute("""
                UPDATE alumnos 
                SET nombre=%s, carrera=%s 
                WHERE matricula=%s
            """, (nombre, carrera, matricula))

            # ✅ solo actualizar password si escriben algo
            if nueva_password:
                password_hash = generate_password_hash(nueva_password)

                cursor.execute("""
                    UPDATE usuarios 
                    SET password=%s 
                    WHERE usuario=%s
                """, (password_hash, matricula))

            conn.commit()
            conn.close()

            return redirect('/admin_alumnos')

        except Exception as e:
            conn.close()
            return f"❌ Error: {str(e)}"

    # ✅ GET
    cursor.execute("""
        SELECT matricula, nombre, carrera 
        FROM alumnos 
        WHERE matricula=%s
    """, (matricula,))
    
    alumno = cursor.fetchone()
    conn.close()

    return render_template('editar_alumno.html', alumno=alumno, mensaje=mensaje)

# ======================
# Historial Alumnos
# ======================



@app.route('/historial_admin', methods=['GET','POST'])
def historial_admin():

    if 'usuario' not in session:
        return redirect('/')

    if session.get('rol') != 'admin':
        return "Acceso denegado"

    conn = get_db()
    cursor = conn.cursor()

    historial = []
    matricula = ""
    creditos = 0  # ✅ contador

    if request.method == 'POST':

        matricula = request.form['matricula']

        cursor.execute("""
        SELECT 
            inscripciones.periodo,
            inscripciones.semestre,
            inscripciones.actividad,
            inscripciones.genero,
            inscripciones.telefono,
            COALESCE(resultados.resultado, 'Sin evaluar')
        FROM inscripciones
        LEFT JOIN resultados
            ON resultados.matricula = inscripciones.matricula
            AND resultados.actividad = inscripciones.actividad
            AND resultados.periodo = inscripciones.periodo
        WHERE inscripciones.matricula=%s
        ORDER BY inscripciones.periodo DESC
        """, (matricula,))

        historial = cursor.fetchall()

        # ✅ contar créditos liberados
        creditos = sum(1 for h in historial if h[5] == "Acreditado")

    conn.close()

    return render_template(
        'historial_admin.html',
        historial=historial,
        matricula=matricula,
        creditos=creditos
    )

# ======================
# DASHBOARD
# ======================

@app.route('/dashboard')
def dashboard():

    if 'usuario' not in session:
        return redirect('/')

    if session.get('rol') != 'admin':
        return "Acceso denegado"

    conn = get_db()
    cursor = conn.cursor()

    # ✅ Total alumnos
    cursor.execute("SELECT COUNT(*) FROM inscripciones")
    total_alumnos = cursor.fetchone()[0]

    # ✅ Total talleres
    cursor.execute("SELECT COUNT(*) FROM actividades")
    total_talleres = cursor.fetchone()[0]

    # ✅ Total docentes
    cursor.execute("SELECT COUNT(*) FROM usuarios WHERE rol='docente'")
    total_docentes = cursor.fetchone()[0]

    # ✅ Periodo activo
    cursor.execute("SELECT periodo FROM periodos WHERE activo=1")
    periodo = cursor.fetchone()
    periodo = periodo[0] if periodo else "No activo"

    # ✅ Alumnos por taller
    cursor.execute("""
    SELECT actividad, COUNT(*) 
    FROM inscripciones
    GROUP BY actividad
    """)
    por_taller = cursor.fetchall()

    conn.close()

    return render_template('dashboard.html',
        total_alumnos=total_alumnos,
        total_talleres=total_talleres,
        total_docentes=total_docentes,
        periodo=periodo,
        por_taller=por_taller
    )


# ======================
# TALLERES
# ======================
@app.route('/admin_talleres')
def admin_talleres():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM actividades")
    talleres = cursor.fetchall()
    conn.close()
    return render_template('admin_talleres.html', talleres=talleres)

# ======================
# nuevos TALLERES
# ======================

@app.route('/nuevo_taller', methods=['GET','POST'])
def nuevo_taller():

    if 'usuario' not in session:
        return redirect('/')

    mensaje = ""

    if request.method == 'POST':

        nombre = request.form['nombre']

        if not nombre.strip():
            mensaje = "❌ Debes escribir un nombre válido"
            return render_template('nuevo_taller.html', mensaje=mensaje)

        nombre = nombre.strip().title()

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT * FROM actividades 
            WHERE LOWER(nombre) = LOWER(%s)
        """, (nombre,))
        
        if cursor.fetchone():
            conn.close()
            mensaje = "❌ El taller ya existe"
            return render_template('nuevo_taller.html', mensaje=mensaje)

        # ✅ 👇 AQUÍ VA LA CORRECCIÓN IMPORTANTE
        docente = session['usuario']

        cursor.execute("""
        INSERT INTO actividades (nombre, docente)
        VALUES (%s, %s)
        """, (nombre, docente))

        conn.commit()
        conn.close()

        mensaje = "✅ Taller registrado correctamente"
        return render_template('nuevo_taller.html', mensaje=mensaje)

    return render_template('nuevo_taller.html')


# ======================
# Editar TALLERES
# ======================
@app.route('/editar_taller/<int:id>', methods=['GET','POST'])
def editar_taller(id):

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'POST':

        nombre = request.form['nombre'].strip().title()

        if not nombre:
            conn.close()
            return render_template('editar_taller.html', mensaje="❌ Nombre inválido")

        # 🔍 Validar duplicado
        cursor.execute("""
        SELECT * FROM actividades 
        WHERE LOWER(nombre) = LOWER(%s) AND id != %s
        """, (nombre, id))

        if cursor.fetchone():
            conn.close()
            return render_template('editar_taller.html', mensaje="❌ Ya existe ese taller")

        # ✅ Actualizar
        cursor.execute("UPDATE actividades SET nombre=%s WHERE id=%s", (nombre, id))

        conn.commit()
        conn.close()

        return redirect('/admin_talleres')

    # 🔹 Obtener datos actuales
    cursor.execute("SELECT * FROM actividades WHERE id=%s", (id,))
    taller = cursor.fetchone()

    conn.close()

    return render_template('editar_taller.html', taller=taller)


# ======================
# Eliminar TALLERES
# ======================


@app.route('/eliminar_taller/<int:id>')
def eliminar_taller(id):

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("DELETE FROM actividades WHERE id=%s", (id,))

    conn.commit()
    conn.close()

    return redirect('/admin_talleres')


# ======================
# PERIODOS
# ======================
@app.route('/admin_periodos')
def admin_periodos():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM periodos")
    periodos = cursor.fetchall()
    conn.close()
    return render_template('admin_periodos.html', periodos=periodos)

# ======================
# activar PERIODOS
# ======================

@app.route('/activar_periodo/<periodo>')
def activar_periodo(periodo):

    if 'usuario' not in session:
        return redirect('/')

    if session.get('rol') != 'admin':
        return "Acceso denegado"

    conn = get_db()
    cursor = conn.cursor()

    # ✅ 1. Desactivar todos
    cursor.execute("UPDATE periodos SET activo = 0")

    # ✅ 2. Activar el seleccionado
    cursor.execute("UPDATE periodos SET activo = 1 WHERE periodo = %s", (periodo,))

    conn.commit()
    conn.close()

    return redirect('/admin_periodos')




# ======================
# EXCEL (INTELIGENTE)
# ======================
@app.route('/cargar_excel', methods=['GET','POST'])
def cargar_excel():

    if session.get('rol') != 'admin':
        return "Acceso denegado"

    mensaje = ""

    if request.method == 'POST':

        archivo = request.files.get('archivo')

        if not archivo or archivo.filename == '':
            return render_template("cargar_excel.html",
                                   mensaje="❌ Debes seleccionar un archivo")

        try:
            df = pd.read_excel(archivo)

            # ✅ normalizar columnas
            df.columns = df.columns.str.strip().str.lower()

            conn = get_db()
            cursor = conn.cursor()

            nuevos = 0
            existentes = 0
            errores = 0

            for i, row in df.iterrows():

                try:
                    matricula = str(row.get('matricula', '')).strip()
                    nombre = str(row.get('nombre', '')).strip()
                    carrera = str(row.get('carrera', '')).strip()

                    # ✅ validación básica
                    if not matricula or not nombre:
                        errores += 1
                        continue

                    # 🔍 verificar alumno existente
                    cursor.execute("SELECT 1 FROM alumnos WHERE matricula=%s", (matricula,))
                    existe_alumno = cursor.fetchone()

                    if existe_alumno:
                        existentes += 1
                        continue

                    # ✅ insertar alumno
                    cursor.execute("""
                        INSERT INTO alumnos (matricula, nombre, carrera)
                        VALUES (%s, %s, %s)
                    """, (matricula, nombre, carrera))

                    # 🔍 verificar usuario existente
                    cursor.execute("SELECT 1 FROM usuarios WHERE usuario=%s", (matricula,))
                    existe_usuario = cursor.fetchone()

                    if not existe_usuario:
                        password_hash = generate_password_hash(matricula)

                        cursor.execute("""
                            INSERT INTO usuarios (usuario, password, rol, cambio_password)
                            VALUES (%s, %s, 'alumno', 0)
                        """, (matricula, password_hash))

                    nuevos += 1

                except Exception as e:
                    print(f"❌ Error fila {i}: {e}")
                    errores += 1

            conn.commit()
            conn.close()

            mensaje = f"""
✅ Nuevos: {nuevos}
⚠️ Existentes: {existentes}
❌ Errores: {errores}
"""

        except Exception as e:
            mensaje = f"❌ Error al procesar archivo: {str(e)}"

        return render_template("cargar_excel.html", mensaje=mensaje)

    return render_template("cargar_excel.html")

# ======================
# REPORTES ✅
# ======================

@app.route('/reportes', methods=['GET','POST'])
def reportes():

    if session.get('rol') != 'admin':
        return "Acceso denegado"

    conn = get_db()
    cursor = conn.cursor()

    # 🔹 Obtener lista de periodos
    cursor.execute("SELECT periodo FROM periodos")
    lista_periodos = cursor.fetchall()

    periodo_seleccionado = None

    # Si el usuario selecciona uno
    if request.method == 'POST':
        periodo_seleccionado = request.form['periodo']
    else:
        # por default el activo
        cursor.execute("SELECT periodo FROM periodos WHERE activo=1")
        periodo_seleccionado = cursor.fetchone()[0]

    # ✅ CONSULTAS FILTRADAS
    cursor.execute("SELECT COUNT(*) FROM inscripciones WHERE periodo=%s", (periodo_seleccionado,))
    total = cursor.fetchone()[0]

    cursor.execute("""
    SELECT actividad, COUNT(*) 
    FROM inscripciones
    WHERE periodo=%s
    GROUP BY actividad
    """, (periodo_seleccionado,))
    por_taller = cursor.fetchall()

    cursor.execute("""
    SELECT genero, COUNT(*) 
    FROM inscripciones
    WHERE periodo=%s
    GROUP BY genero
    """, (periodo_seleccionado,))
    por_genero = cursor.fetchall()

    cursor.execute("""
    SELECT semestre, COUNT(*) 
    FROM inscripciones
    WHERE periodo=%s
    GROUP BY semestre
    """, (periodo_seleccionado,))
    por_semestre = cursor.fetchall()

    cursor.execute("""
    SELECT alumnos.carrera, COUNT(*) 
    FROM inscripciones
    JOIN alumnos ON alumnos.matricula = inscripciones.matricula
    WHERE periodo=%s
    GROUP BY alumnos.carrera
    """, (periodo_seleccionado,))
    por_carrera = cursor.fetchall()

    cursor.execute("""
    SELECT actividad, genero, COUNT(*) 
    FROM inscripciones
    WHERE periodo=%s
    GROUP BY actividad, genero
    """, (periodo_seleccionado,))
    por_taller_genero = cursor.fetchall()

    conn.close()

    return render_template("reportes.html",
        total=total,
        por_taller=por_taller,
        por_genero=por_genero,
        por_semestre=por_semestre,
        por_carrera=por_carrera,
        por_taller_genero=por_taller_genero,
        lista_periodos=lista_periodos,
        periodo_seleccionado=periodo_seleccionado
    )

# ======================
# ALUMNO
# ======================
@app.route('/alumno')
def alumno():

    if 'usuario' not in session:
        return redirect('/')

    if session.get('rol') != 'alumno':
        return "Acceso denegado"

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT nombre, carrera FROM alumnos WHERE matricula=%s",
                   (session['usuario'],))
    alumno = cursor.fetchone()

    cursor.execute("SELECT periodo FROM periodos WHERE activo=1")
    periodo = cursor.fetchone()[0]

    cursor.execute("SELECT nombre FROM actividades")
    talleres = cursor.fetchall()

    conn.close()

    estado = obtener_estado_fechas()

    return render_template('alumno.html',
        nombre=alumno[0],
        carrera=alumno[1],
        matricula=session['usuario'],
        periodo=periodo,
        talleres=talleres,
        estado=estado
    )




# ======================
# REGISTRO de alumno en taller
# ======================
@app.route('/registrar_taller', methods=['POST'])
def registrar_taller():

    estado = obtener_estado_fechas()

    if estado != "activo":
        return "❌ El periodo de inscripción está cerrado"

    telefono = request.form['telefono']
    genero = request.form['genero']

    if not telefono.isdigit() or len(telefono) != 10:
        return "❌ Teléfono inválido"

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT periodo FROM periodos WHERE activo=1")
    periodo = cursor.fetchone()[0]

    cursor.execute("""
    SELECT * FROM inscripciones 
    WHERE matricula=%s AND periodo=%s
    """, (session['usuario'], periodo))

    if cursor.fetchone():
        return "❌ Ya registrado"

    cursor.execute("""
    INSERT INTO inscripciones
    (matricula, semestre, genero, telefono, actividad, periodo)
    VALUES (%s,%s,%s,%s,%s,%s)
    """,(
        session['usuario'],
        request.form['semestre'],
        genero,
        telefono,
        request.form['actividad'],
        periodo
    ))

    conn.commit()
    conn.close()

    return redirect('/alumno')

# ======================
# Nuevo periodo escolar
# ======================
@app.route('/nuevo_periodo', methods=['GET','POST'])
def nuevo_periodo():

    if request.method == 'POST':

        periodo = request.form['periodo']

        if not periodo:
            return "❌ Debes escribir un periodo"

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO periodos (periodo, activo) 
        VALUES (%s, 0)
        """, (periodo,))

        conn.commit()
        conn.close()

        return redirect('/admin_periodos')

    return render_template('nuevo_periodo.html')






# ======================
# LOGOUT
# ======================
@app.route('/logout')
def logout():
    session.clear()
    return redirect('/')

init_db()
cargar_datos()

if __name__ == '__main__':
    app.run()
